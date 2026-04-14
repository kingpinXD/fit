#!/usr/bin/env python3
"""Parse workout programme XLSX files into standardized JSON for the Fit app."""

import json
import re
import sys
from datetime import datetime

import openpyxl


PROGRAMME_NAME = "Jeff Nippard's Essentials Program"
WEEK_PATTERN = re.compile(r"^Week\s+(\d+)")
SKIP_VALUES = {"Exercise", "Warm-up Sets", "Working Sets", "Reps", "Load", "RPE", "Rest",
               "Notes", "Warm-up Sets (see page 15 for details)", "Substitution Option 1",
               "Substitution Option 2"}


def parse_rpe(value) -> str:
    """Parse RPE from a cell value.

    Excel sometimes stores RPE ranges like "8-9" as dates (2022-08-09).
    Detect datetimes and extract month-day as the RPE range.
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"{value.month}-{value.day}"
    if isinstance(value, (int, float)):
        return str(int(value))
    return str(value).strip()


def parse_warmup(value) -> str:
    """Parse warmup sets — same datetime quirk as RPE."""
    if value is None:
        return "0"
    if isinstance(value, datetime):
        return f"{value.month}-{value.day}"
    if isinstance(value, (int, float)):
        return str(int(value))
    return str(value).strip()


def make_exercise(name: str, warmup, sets: int, reps, rpe, rest, notes, order: int,
                   sub1="", sub2="", video_url="", sub1_video_url="", sub2_video_url="") -> dict:
    return {
        "name": str(name).strip(),
        "warmupSets": parse_warmup(warmup),
        "sets": int(sets),
        "reps": str(reps).strip() if reps else "",
        "rpe": parse_rpe(rpe),
        "rest": str(rest).strip() if rest else "",
        "notes": str(notes).strip() if notes else "",
        "order": order,
        "sub1": str(sub1).strip() if sub1 else "",
        "sub2": str(sub2).strip() if sub2 else "",
        "videoUrl": str(video_url).strip() if video_url else "",
        "sub1VideoUrl": str(sub1_video_url).strip() if sub1_video_url else "",
        "sub2VideoUrl": str(sub2_video_url).strip() if sub2_video_url else "",
    }


def parse_programme(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    weeks = []
    current_week = None
    current_day = None
    # Track day name occurrences within a week to disambiguate duplicates
    # (e.g., 4x has Upper, Lower, Upper, Lower → Upper, Lower, Upper #2, Lower #2)
    day_counts_per_week = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=12):
        col_b = row[0].value   # Column B
        col_c = row[1].value   # Column C
        col_d = row[2].value   # Column D (warmup sets)
        col_e = row[3].value   # Column E (working sets)
        col_f = row[4].value   # Column F (reps)
        col_h = row[6].value   # Column H (RPE)
        col_i = row[7].value   # Column I (rest)
        col_j = row[8].value   # Column J (sub1)
        col_k = row[9].value   # Column K (sub2)
        col_l = row[10].value  # Column L (notes)
        # Hyperlinks
        video_url = row[1].hyperlink.target if row[1].hyperlink else ""
        sub1_video_url = row[8].hyperlink.target if row[8].hyperlink else ""
        sub2_video_url = row[9].hyperlink.target if row[9].hyperlink else ""

        if col_b and isinstance(col_b, str):
            col_b = col_b.strip()

            # Week header row (e.g., "Week 1")
            week_match = WEEK_PATTERN.match(col_b)
            if week_match:
                current_week = {"week": int(week_match.group(1)), "days": []}
                weeks.append(current_week)
                current_day = None
                day_counts_per_week = {}
                continue

            # Skip header rows and rest days
            if col_b in SKIP_VALUES or "Suggested" in col_b or "Copyright" in col_b:
                continue

            # Day label row — any non-week, non-skip value in column B with an exercise in C
            if current_week is not None:
                count = day_counts_per_week.get(col_b, 0) + 1
                day_counts_per_week[col_b] = count
                unique_day_name = f"{col_b} #{count}" if count > 1 else col_b
                current_day = {"day": unique_day_name, "exercises": []}
                current_week["days"].append(current_day)
                if col_c and col_e is not None:
                    current_day["exercises"].append(
                        make_exercise(col_c, col_d, col_e, col_f, col_h, col_i, col_l, order=1,
                                      sub1=col_j, sub2=col_k, video_url=video_url,
                                      sub1_video_url=sub1_video_url, sub2_video_url=sub2_video_url)
                    )
                continue

        # Exercise row: no day/week marker in B, but has exercise name in C
        if col_c and col_e is not None and current_day is not None:
            order = len(current_day["exercises"]) + 1
            current_day["exercises"].append(
                make_exercise(col_c, col_d, col_e, col_f, col_h, col_i, col_l, order=order,
                              sub1=col_j, sub2=col_k, video_url=video_url,
                              sub1_video_url=sub1_video_url, sub2_video_url=sub2_video_url)
            )

    return {"name": PROGRAMME_NAME, "weeks": weeks}


def main():
    if len(sys.argv) != 3:
        print(f"Usage: {sys.argv[0]} <input.xlsx> <output.json>", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    json_path = sys.argv[2]

    programme = parse_programme(xlsx_path)

    with open(json_path, "w") as f:
        json.dump(programme, f, indent=2)

    # Summary
    total_weeks = len(programme["weeks"])
    for w in programme["weeks"]:
        days_summary = ", ".join(
            f"{d['day']}({len(d['exercises'])})" for d in w["days"]
        )
        print(f"  Week {w['week']}: {len(w['days'])} days — {days_summary}")
    print(f"\nWrote {total_weeks} weeks to {json_path}")


if __name__ == "__main__":
    main()
